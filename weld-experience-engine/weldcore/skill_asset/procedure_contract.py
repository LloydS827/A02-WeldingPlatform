from __future__ import annotations

from collections import Counter
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook


DEFAULT_PROCEDURE_WORKBOOK_PATH = (
    Path(__file__).resolve().parents[3] / "docs" / "焊接工艺数据库主要参数表.xlsx"
)
CONTRACT_VERSION = "k01.v0.1"
MAIN_SHEET_NAME = "工艺数据库参数总表"
REQUIRED_HEADERS = ("参数类别", "参数名称", "参数说明", "数据类型", "是否必填", "备注")

_REQUIREMENT_LEVELS = {
    "必填": "required",
    "条件必填": "conditional_required",
    "可选": "supplemental",
}
_DATA_TYPES = {
    "文本": "text",
    "数值": "number",
    "整数": "integer",
}
_EXPECTED_ROW_COUNT = 48
_EXPECTED_FIELD_COUNT = 47
_EXPECTED_CATEGORY_COUNT = 8
_EXPECTED_REQUIREMENT_SUMMARY = {
    "required": 21,
    "conditional_required": 12,
    "supplemental": 14,
}
_EXPECTED_DATA_TYPE_SUMMARY = {
    "text": 25,
    "number": 21,
    "integer": 1,
}
_FIELD_OVERRIDES: dict[str, dict[str, Any]] = {
    "WPS编号": {
        "field_id": "wps_number",
        "acquisition_mode": "human_required",
        "human_role": "welding_procedure_engineer",
        "a02_target_path": "ExpertReviewRecord.required_real_context",
        "nv01_usage": ["procedure_gate", "expert_gate"],
        "blocks": ["expert_review", "wps_pqr_release"],
    },
    "PQR编号": {
        "field_id": "pqr_number",
        "acquisition_mode": "human_required",
        "human_role": "welding_procedure_engineer",
        "a02_target_path": "ExpertReviewRecord.required_real_context",
        "nv01_usage": ["procedure_gate"],
        "blocks": ["wps_pqr_release"],
    },
    "热输入(kJ/mm)": {
        "field_id": "heat_input_kj_per_mm",
        "acquisition_mode": "system_computed",
        "human_role": "welding_procedure_engineer_review",
        "a02_target_path": "ManipulationSkillAsset.constraints.process_parameters.heat_input",
        "nv01_usage": ["training_readiness_report", "domain_randomization_recipe"],
        "blocks": ["wps_pqr_release"],
    },
    "焊接速度(mm/min)": {
        "field_id": "travel_speed_mm_per_min",
        "acquisition_mode": "asset_or_simulation_inferred",
        "human_role": "welding_robotics_engineer_review",
        "a02_target_path": "ManipulationSkillAsset.motion.tcp_trajectory",
        "nv01_usage": ["isaac_replay_config", "domain_randomization_recipe"],
        "blocks": ["expert_review"],
    },
    "坡口角度α(°)": {
        "field_id": "groove_angle_deg",
        "acquisition_mode": "human_confirmed_or_imported",
        "human_role": "welding_procedure_engineer",
        "a02_target_path": "ManipulationSkillAsset.constraints.joint_geometry.groove_angle",
        "nv01_usage": ["OpenUSD process_metadata", "domain_randomization_recipe"],
        "blocks": ["expert_review"],
    },
    "根部间隙R(mm)": {
        "field_id": "root_gap_mm",
        "acquisition_mode": "human_confirmed_or_imported",
        "human_role": "welding_procedure_engineer",
        "a02_target_path": "ManipulationSkillAsset.constraints.joint_geometry.root_gap",
        "nv01_usage": ["OpenUSD process_metadata", "domain_randomization_recipe"],
        "blocks": ["expert_review"],
    },
    "焊接电流(A)": {
        "field_id": "welding_current_a",
        "acquisition_mode": "workcell_logged",
        "human_role": "welding_procedure_engineer_review",
        "a02_target_path": "ManipulationSkillAsset.constraints.process_parameters.current",
        "nv01_usage": ["procedure_parameter_inputs", "domain_randomization_recipe"],
        "blocks": ["expert_review"],
    },
    "焊接电压(V)": {
        "field_id": "welding_voltage_v",
        "acquisition_mode": "workcell_logged",
        "human_role": "welding_procedure_engineer_review",
        "a02_target_path": "ManipulationSkillAsset.constraints.process_parameters.voltage",
        "nv01_usage": ["procedure_parameter_inputs", "domain_randomization_recipe"],
        "blocks": ["expert_review"],
    },
    "无损检测等级": {
        "field_id": "ndt_acceptance_level",
        "acquisition_mode": "human_required",
        "human_role": "quality_engineer",
        "a02_target_path": "ExpertReviewRecord.required_real_context",
        "nv01_usage": ["expert_gate", "training_readiness_report"],
        "blocks": ["expert_review", "wps_pqr_release"],
    },
}


def build_weld_procedure_knowledge_contract(
    workbook_path: str | Path = DEFAULT_PROCEDURE_WORKBOOK_PATH,
) -> dict[str, Any]:
    path = Path(workbook_path)
    if not path.exists():
        raise FileNotFoundError(path)

    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook[MAIN_SHEET_NAME]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = tuple(rows[0][: len(REQUIRED_HEADERS)])
    if headers != REQUIRED_HEADERS:
        raise ValueError(f"Unexpected procedure workbook headers: {headers!r}")

    fields: list[dict[str, Any]] = []
    categories: list[str] = []
    current_category = ""
    for index, row in enumerate(rows[1:], start=1):
        category, display_name, description, data_type, requirement, remark = row[:6]
        if category:
            current_category = str(category)
            categories.append(current_category)
        if not display_name:
            continue
        fields.append(
            _build_field(
                index=index,
                category=current_category,
                display_name=str(display_name),
                description=str(description or ""),
                data_type=str(data_type),
                requirement=str(requirement),
                remark=str(remark or ""),
            )
        )

    requirement_counts = Counter(field["requirement_level"] for field in fields)
    requirement_summary = {
        "required": requirement_counts["required"],
        "conditional_required": requirement_counts["conditional_required"],
        "supplemental": requirement_counts["supplemental"],
    }
    data_type_counts = Counter(field["data_type"] for field in fields)
    data_type_summary = {
        "text": data_type_counts["text"],
        "number": data_type_counts["number"],
        "integer": data_type_counts["integer"],
    }
    _validate_baseline(
        row_count=len(rows),
        fields=fields,
        categories=categories,
        requirement_summary=requirement_summary,
        data_type_summary=data_type_summary,
    )
    return {
        "source_workbook_ref": str(path.resolve()),
        "contract_version": CONTRACT_VERSION,
        "row_count": len(rows),
        "field_count": len(fields),
        "category_count": len(categories),
        "requirement_summary": requirement_summary,
        "data_type_summary": data_type_summary,
        "categories": categories,
        "fields": fields,
        "a02_target_paths": sorted(
            {field["a02_target_path"] for field in fields if field["a02_target_path"]}
        ),
        "nv01_usage_tags": sorted(
            {usage for field in fields for usage in field["nv01_usage"]}
        ),
        "evidence_boundary": [
            "excel_field_contract_source",
            "not_formal_WPS_PQR",
            "requires_human_confirmation_before_expert_review",
        ],
    }


def build_weld_procedure_parameter_set(
    skill_asset: Any,
    contract: dict[str, Any],
    artifact_refs: dict[str, str] | None = None,
) -> dict[str, Any]:
    values: dict[str, dict[str, Any]] = {}
    missing_required_fields: list[str] = []
    missing_conditional_fields: list[str] = []
    supplemental_gaps: list[str] = []
    computed_fields: list[str] = []
    blocked_fields: list[str] = []
    blocked_computed_fields: list[str] = []
    inferred_fields: list[str] = []
    workcell_logged_gaps: list[str] = []

    for field in contract["fields"]:
        field_id = field["field_id"]
        status = _parameter_status_for_field(field, skill_asset)
        values[field_id] = status

        if status["coverage_status"] == "missing_required":
            missing_required_fields.append(field_id)
        if status["coverage_status"] == "missing_conditional":
            missing_conditional_fields.append(field_id)
        if status["coverage_status"] == "supplemental_gap":
            supplemental_gaps.append(field_id)
        if status["coverage_status"].startswith("blocked_"):
            blocked_fields.append(field_id)
        if field["acquisition_mode"] == "system_computed":
            computed_fields.append(field_id)
            if status["coverage_status"].startswith("blocked_"):
                blocked_computed_fields.append(field_id)
        if field["acquisition_mode"] == "asset_or_simulation_inferred":
            inferred_fields.append(field_id)
        if field["acquisition_mode"] == "workcell_logged" and status["value"] is None:
            workcell_logged_gaps.append(field_id)

    return {
        "parameter_set_id": f"procedure-params-{skill_asset.asset_id}",
        "skill_asset_id": skill_asset.asset_id,
        "contract_version": contract["contract_version"],
        "values": values,
        "missing_required_fields": sorted(missing_required_fields),
        "missing_conditional_fields": sorted(missing_conditional_fields),
        "supplemental_gaps": sorted(supplemental_gaps),
        "computed_fields": sorted(computed_fields),
        "blocked_fields": sorted(blocked_fields),
        "blocked_computed_fields": sorted(blocked_computed_fields),
        "inferred_fields": sorted(inferred_fields),
        "workcell_logged_gaps": sorted(workcell_logged_gaps),
        "source_summary": {
            "skill_asset_source_type": skill_asset.source_type,
            "skill_asset_id": skill_asset.asset_id,
            "artifact_refs": artifact_refs or {},
        },
        "review_boundary": [
            "not_formal_WPS_PQR",
            "missing_human_required_fields",
            "simulation_only_values_require_expert_review",
        ],
    }


def build_weld_procedure_validation_report(
    parameter_set: dict[str, Any],
    contract: dict[str, Any],
) -> dict[str, Any]:
    human_required_modes = {"human_required", "human_confirmed_or_imported"}
    human_required_gaps = [
        field_id
        for field_id in parameter_set["missing_required_fields"]
        if _field_by_id(contract, field_id)["acquisition_mode"] in human_required_modes
    ]
    workcell_logged_gaps = [
        field_id
        for field_id in parameter_set["missing_required_fields"]
        if _field_by_id(contract, field_id)["acquisition_mode"] == "workcell_logged"
    ]

    not_ready_reasons: list[str] = []
    if human_required_gaps:
        not_ready_reasons.append("blocked_by_missing_human_required_fields")
    if parameter_set["missing_conditional_fields"]:
        not_ready_reasons.append("blocked_by_missing_conditional_procedure_fields")
    if workcell_logged_gaps:
        not_ready_reasons.append("blocked_by_missing_workcell_logged_fields")
    if parameter_set.get("blocked_computed_fields"):
        not_ready_reasons.append("blocked_by_missing_real_process_inputs")

    return {
        "validation_status": not_ready_reasons[0]
        if not_ready_reasons
        else "ready_for_procedure_contract_review",
        "ready_for_procedure_contract_review": True,
        "ready_for_expert_review": not not_ready_reasons,
        "ready_for_simulation_replay_package_design": True,
        "ready_for_training_design_review": True,
        "readiness_scope": "design/review draft readiness only",
        "readiness_notes": [
            "not_formal_WPS_PQR",
            "not_expert_approved",
            "not_isaac_runtime_replay_ready",
            "not_policy_training_ready",
        ],
        "not_ready_reasons": not_ready_reasons,
        "field_coverage": _field_coverage(parameter_set, contract),
        "human_required_gaps": sorted(human_required_gaps),
        "computed_fields": parameter_set["computed_fields"],
        "blocked_fields": parameter_set.get("blocked_fields", []),
        "blocked_computed_fields": parameter_set.get("blocked_computed_fields", []),
        "inferred_fields": parameter_set["inferred_fields"],
        "workcell_logged_gaps": sorted(workcell_logged_gaps),
        "wps_pqr_boundary": "not_formal_WPS_PQR",
    }


def build_procedure_to_nv01_mapping_matrix(contract: dict[str, Any]) -> dict[str, Any]:
    return {
        "matrix_id": f"procedure-to-nv01-{contract['contract_version']}",
        "contract_version": contract["contract_version"],
        "field_count": contract["field_count"],
        "field_mappings": {
            field["field_id"]: {
                "display_name": field["display_name"],
                "requirement_level": field["requirement_level"],
                "acquisition_mode": field["acquisition_mode"],
                "a02_targets": _a02_targets_for_field(field),
                "nv01_targets": _nv01_targets_for_field(field),
                "blocks": field["blocks"],
                "evidence_boundary": field["evidence_boundary"],
            }
            for field in contract["fields"]
        },
    }


def _build_field(
    *,
    index: int,
    category: str,
    display_name: str,
    description: str,
    data_type: str,
    requirement: str,
    remark: str,
) -> dict[str, Any]:
    override = _FIELD_OVERRIDES.get(display_name, {})
    requirement_level = _map_requirement(requirement)
    field = {
        "field_id": f"field_{index:03d}",
        "category": category,
        "display_name": display_name,
        "description": description,
        "data_type": _map_data_type(data_type),
        "unit": _parse_unit(display_name),
        "requirement_level": requirement_level,
        "required_when": remark if requirement_level == "conditional_required" else "",
        "acquisition_mode": "human_confirmed_or_imported",
        "human_role": "",
        "a02_target_path": "",
        "nv01_usage": [],
        "blocks": [],
        "evidence_boundary": "excel_contract_field",
    }
    field.update(override)
    return field


def _parameter_status_for_field(field: dict[str, Any], skill_asset: Any) -> dict[str, Any]:
    acquisition_mode = field["acquisition_mode"]
    value = None
    source = None
    evidence_boundary = ["not_formal_WPS_PQR"]
    coverage_status = _missing_status_for_requirement(field)

    if field["field_id"] == "travel_speed_mm_per_min":
        value = _infer_travel_speed_mm_per_min(skill_asset)
        source = "ManipulationSkillAsset.motion.tcp_trajectory"
        evidence_boundary = [
            "simulation_inferred_not_wps_validated",
            "not_workcell_logged",
            "not_formal_WPS_PQR",
        ]
        coverage_status = (
            "simulation_inferred_candidate" if value is not None else coverage_status
        )
    elif acquisition_mode == "system_computed":
        coverage_status = "blocked_missing_real_process_inputs"
        source = (
            "requires_real_welding_current_a_welding_voltage_v_"
            "and_travel_speed_mm_per_min"
        )
        evidence_boundary = [
            "computed_value_blocked_until_real_process_inputs",
            "not_formal_WPS_PQR",
        ]

    return {
        "field_id": field["field_id"],
        "display_name": field["display_name"],
        "value": value,
        "coverage_status": coverage_status,
        "acquisition_mode": acquisition_mode,
        "source": source,
        "evidence_boundary": evidence_boundary,
    }


def _missing_status_for_requirement(field: dict[str, Any]) -> str:
    if field["requirement_level"] == "required":
        return "missing_required"
    if field["requirement_level"] == "conditional_required":
        return "missing_conditional"
    return "supplemental_gap"


def _infer_travel_speed_mm_per_min(skill_asset: Any) -> float | None:
    trajectory = list(getattr(skill_asset, "motion", {}).get("tcp_trajectory") or [])
    if len(trajectory) < 2:
        return None

    first = trajectory[0]
    last = trajectory[-1]
    duration_s = float(last["t"]) - float(first["t"])
    if duration_s <= 0:
        return None

    path_length = 0.0
    previous = first
    for point in trajectory[1:]:
        dx = float(point["x"]) - float(previous["x"])
        dy = float(point["y"]) - float(previous["y"])
        dz = float(point["z"]) - float(previous["z"])
        path_length += (dx * dx + dy * dy + dz * dz) ** 0.5
        previous = point

    return round(path_length * 1000.0 * 60.0 / duration_s, 3)


def _field_by_id(contract: dict[str, Any], field_id: str) -> dict[str, Any]:
    for field in contract["fields"]:
        if field["field_id"] == field_id:
            return field
    raise KeyError(field_id)


def _field_coverage(
    parameter_set: dict[str, Any],
    contract: dict[str, Any],
) -> dict[str, Any]:
    statuses = Counter(
        parameter_set["values"][field["field_id"]]["coverage_status"]
        for field in contract["fields"]
    )
    return {
        "field_count": contract["field_count"],
        "by_status": dict(sorted(statuses.items())),
        "covered_field_count": sum(
            count
            for status, count in statuses.items()
            if status
            not in {"missing_required", "missing_conditional", "supplemental_gap"}
            and not status.startswith("blocked_")
        ),
    }


def _a02_targets_for_field(field: dict[str, Any]) -> list[str]:
    return [field["a02_target_path"]] if field["a02_target_path"] else []


def _nv01_targets_for_field(field: dict[str, Any]) -> list[str]:
    targets = [_human_readable_nv01_target(tag) for tag in field["nv01_usage"]]
    return sorted({target for target in targets if target})


def _human_readable_nv01_target(tag: str) -> str:
    target_by_tag = {
        "isaac_replay_config": "Isaac Sim replay config",
        "procedure_parameter_inputs": "Isaac Sim replay config",
        "procedure_gate": "procedure contract gate",
        "expert_gate": "ExpertReviewRecord.required_real_context",
        "OpenUSD process_metadata": "OpenUSD process_metadata",
        "domain_randomization_recipe": "domain_randomization_recipe",
        "training_readiness_report": "training_readiness_report",
    }
    return target_by_tag.get(tag, tag)


def _validate_baseline(
    *,
    row_count: int,
    fields: list[dict[str, Any]],
    categories: list[str],
    requirement_summary: dict[str, int],
    data_type_summary: dict[str, int],
) -> None:
    if (
        row_count != _EXPECTED_ROW_COUNT
        or len(fields) != _EXPECTED_FIELD_COUNT
        or len(categories) != _EXPECTED_CATEGORY_COUNT
        or requirement_summary != _EXPECTED_REQUIREMENT_SUMMARY
        or data_type_summary != _EXPECTED_DATA_TYPE_SUMMARY
    ):
        raise ValueError(
            "Weld procedure workbook baseline drift: "
            f"rows={row_count}, fields={len(fields)}, categories={len(categories)}, "
            f"requirement_summary={requirement_summary}, "
            f"data_type_summary={data_type_summary}"
        )


def _map_requirement(requirement: str) -> str:
    return _REQUIREMENT_LEVELS[requirement]


def _map_data_type(data_type: str) -> str:
    return _DATA_TYPES[data_type]


def _parse_unit(display_name: str) -> str:
    match = re.search(r"\(([^()]+)\)$", display_name)
    return match.group(1) if match else ""
