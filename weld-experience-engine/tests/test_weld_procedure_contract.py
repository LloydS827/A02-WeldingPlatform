from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from weldcore.simulation_bakeoff import (
    build_simulation_evidence_bundle,
    default_simulation_task_specs,
    run_simlite_reference,
)
from weldcore.skill_asset.builders import (
    build_manipulation_skill_asset_from_simulation_bundle,
)
from weldcore.skill_asset.procedure_contract import (
    DEFAULT_PROCEDURE_WORKBOOK_PATH,
    build_procedure_to_nv01_mapping_matrix,
    build_weld_procedure_knowledge_contract,
    build_weld_procedure_parameter_set,
    build_weld_procedure_validation_report,
)


def _field_by_name(contract, display_name):
    return next(
        field for field in contract["fields"] if field["display_name"] == display_name
    )


def _default_skill_asset():
    task = default_simulation_task_specs()[0]
    bundle = build_simulation_evidence_bundle(task, run_simlite_reference(task))
    return build_manipulation_skill_asset_from_simulation_bundle(bundle)


def test_build_weld_procedure_contract_summarizes_excel_contract():
    contract = build_weld_procedure_knowledge_contract()

    assert contract["source_workbook_ref"].endswith(
        "docs/焊接工艺数据库主要参数表.xlsx"
    )
    assert contract["contract_version"] == "k01.v0.1"
    assert contract["row_count"] == 48
    assert contract["field_count"] == 47
    assert contract["category_count"] == 8
    assert contract["requirement_summary"] == {
        "required": 21,
        "conditional_required": 12,
        "supplemental": 14,
    }
    assert contract["data_type_summary"] == {
        "text": 25,
        "number": 21,
        "integer": 1,
    }
    assert contract["evidence_boundary"] == [
        "excel_field_contract_source",
        "not_formal_WPS_PQR",
        "requires_human_confirmation_before_expert_review",
    ]


def test_build_weld_procedure_contract_preserves_expected_categories():
    contract = build_weld_procedure_knowledge_contract()

    assert contract["categories"] == [
        "母材信息",
        "焊材信息",
        "接头形式",
        "焊接方法",
        "焊接参数",
        "气体参数",
        "质量要求",
        "工艺规程关联",
    ]
    assert all(field["category"] for field in contract["fields"])


def test_build_weld_procedure_contract_maps_representative_fields():
    contract = build_weld_procedure_knowledge_contract()

    base_thickness = _field_by_name(contract, "母材厚度(mm)")
    assert base_thickness["unit"] == "mm"
    assert base_thickness["data_type"] == "number"
    assert base_thickness["requirement_level"] == "required"

    heat_input = _field_by_name(contract, "热输入(kJ/mm)")
    assert heat_input["field_id"] == "heat_input_kj_per_mm"
    assert heat_input["acquisition_mode"] == "system_computed"
    assert "wps_pqr_release" in heat_input["blocks"]

    wps_number = _field_by_name(contract, "WPS编号")
    assert wps_number["acquisition_mode"] == "human_required"
    assert "expert_review" in wps_number["blocks"]

    travel_speed = _field_by_name(contract, "焊接速度(mm/min)")
    assert travel_speed["acquisition_mode"] == "asset_or_simulation_inferred"
    assert "isaac_replay_config" in travel_speed["nv01_usage"]


def test_build_weld_procedure_contract_completes_representative_semantics():
    contract = build_weld_procedure_knowledge_contract()

    pqr_number = _field_by_name(contract, "PQR编号")
    assert pqr_number["human_role"] == "welding_procedure_engineer"
    assert pqr_number["a02_target_path"] == "ExpertReviewRecord.required_real_context"
    assert "procedure_gate" in pqr_number["nv01_usage"]
    assert "wps_pqr_release" in pqr_number["blocks"]

    heat_input = _field_by_name(contract, "热输入(kJ/mm)")
    assert heat_input["human_role"] == "welding_procedure_engineer_review"
    assert (
        heat_input["a02_target_path"]
        == "ManipulationSkillAsset.constraints.process_parameters.heat_input"
    )
    assert "training_readiness_report" in heat_input["nv01_usage"]

    groove_angle = _field_by_name(contract, "坡口角度α(°)")
    assert groove_angle["human_role"] == "welding_procedure_engineer"
    assert groove_angle["a02_target_path"].endswith(".groove_angle")
    assert "OpenUSD process_metadata" in groove_angle["nv01_usage"]
    assert "expert_review" in groove_angle["blocks"]

    welding_current = _field_by_name(contract, "焊接电流(A)")
    assert welding_current["human_role"] == "welding_procedure_engineer_review"
    assert "procedure_parameter_inputs" in welding_current["nv01_usage"]
    assert "expert_review" in welding_current["blocks"]

    ndt_acceptance = _field_by_name(contract, "无损检测等级")
    assert ndt_acceptance["human_role"] == "quality_engineer"
    assert ndt_acceptance["a02_target_path"] == "ExpertReviewRecord.required_real_context"
    assert "expert_gate" in ndt_acceptance["nv01_usage"]
    assert "wps_pqr_release" in ndt_acceptance["blocks"]


def test_build_weld_procedure_contract_raises_for_missing_workbook(tmp_path):
    missing_workbook = tmp_path / "missing.xlsx"

    with pytest.raises(FileNotFoundError):
        build_weld_procedure_knowledge_contract(missing_workbook)


def test_build_weld_procedure_contract_rejects_baseline_drift(tmp_path):
    workbook_path = tmp_path / "drifted.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "工艺数据库参数总表"
    worksheet.append(("参数类别", "参数名称", "参数说明", "数据类型", "是否必填", "备注"))
    worksheet.append(("母材信息", "母材厚度(mm)", "母材的板厚尺寸", "数值", "必填", ""))
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="baseline drift"):
        build_weld_procedure_knowledge_contract(workbook_path)


def test_build_weld_procedure_contract_rejects_data_type_baseline_drift(tmp_path):
    workbook_path = tmp_path / "data-type-drift.xlsx"
    workbook = load_workbook(DEFAULT_PROCEDURE_WORKBOOK_PATH)
    worksheet = workbook["工艺数据库参数总表"]
    worksheet["D3"] = "文本"
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="baseline drift"):
        build_weld_procedure_knowledge_contract(workbook_path)


def test_default_procedure_workbook_path_points_to_repo_docs_workbook():
    assert DEFAULT_PROCEDURE_WORKBOOK_PATH == (
        Path(__file__).resolve().parents[2]
        / "docs"
        / "焊接工艺数据库主要参数表.xlsx"
    )
    assert DEFAULT_PROCEDURE_WORKBOOK_PATH.exists()


def test_parameter_set_marks_missing_human_and_workcell_fields():
    contract = build_weld_procedure_knowledge_contract()
    skill_asset = _default_skill_asset()

    parameter_set = build_weld_procedure_parameter_set(skill_asset, contract)

    assert parameter_set["skill_asset_id"] == skill_asset.asset_id
    assert parameter_set["contract_version"] == contract["contract_version"]
    assert parameter_set["parameter_set_id"].startswith("procedure-params-")
    assert len(parameter_set["values"]) == 47
    assert "wps_number" in parameter_set["missing_required_fields"]
    assert "welding_current_a" in parameter_set["missing_required_fields"]
    assert "heat_input_kj_per_mm" in parameter_set["computed_fields"]
    assert parameter_set["values"]["heat_input_kj_per_mm"]["value"] is None
    assert (
        parameter_set["values"]["heat_input_kj_per_mm"]["coverage_status"]
        == "blocked_missing_real_process_inputs"
    )
    assert "travel_speed_mm_per_min" in parameter_set["inferred_fields"]
    travel_speed = parameter_set["values"]["travel_speed_mm_per_min"]
    assert travel_speed["value"] is not None
    assert (
        "simulation_inferred_not_wps_validated"
        in travel_speed["evidence_boundary"]
    )
    assert "not_formal_WPS_PQR" in parameter_set["review_boundary"]


def test_validation_report_separates_contract_review_from_expert_review():
    contract = build_weld_procedure_knowledge_contract()
    parameter_set = build_weld_procedure_parameter_set(_default_skill_asset(), contract)

    report = build_weld_procedure_validation_report(parameter_set, contract)

    assert report["ready_for_procedure_contract_review"] is True
    assert report["ready_for_simulation_replay_package_design"] is True
    assert report["ready_for_expert_review"] is False
    assert report["validation_status"] == "blocked_by_missing_human_required_fields"
    assert "wps_number" in report["human_required_gaps"]
    assert "welding_current_a" in report["workcell_logged_gaps"]
    assert report["wps_pqr_boundary"] == "not_formal_WPS_PQR"


def test_procedure_to_nv01_mapping_matrix_covers_all_fields():
    contract = build_weld_procedure_knowledge_contract()

    matrix = build_procedure_to_nv01_mapping_matrix(contract)

    assert matrix["contract_version"] == contract["contract_version"]
    assert matrix["field_count"] == 47
    assert len(matrix["field_mappings"]) == 47
    speed = matrix["field_mappings"]["travel_speed_mm_per_min"]
    assert "Isaac Sim replay config" in speed["nv01_targets"]
    assert "domain_randomization_recipe" in speed["nv01_targets"]
    groove = matrix["field_mappings"]["groove_angle_deg"]
    assert "OpenUSD process_metadata" in groove["nv01_targets"]
    wps = matrix["field_mappings"]["wps_number"]
    assert "ExpertReviewRecord.required_real_context" in wps["a02_targets"]
